"""Build BADM 576 lenient grade report.

Reads `badm576_attempts.csv` (dumped from prod DB, covers all three items:
classification quiz, pca quiz, ml_exercise) and produces:

  - Per-item best-attempt-per-student tables (NetID, Date, Raw, Lenient)
  - Combined NetID x item matrix
  - Excel workbook + markdown summary

Lenient grading rules (constructive, not punitive — matches user preference):

  - LLM-graded Classification quiz (max 20):
        lenient = ceil(min(max, raw + 0.20*max))
        i.e. add a 20% engagement bump, capped at max
  - PCA MCQ quiz (max 10): pure auto-grade, identical bump
        lenient = ceil(min(max, raw + 0.20*max))
  - ML Process Exercise (raw is AI score 0-100):
        Already tier-mapped in production: AI > 0 -> Full Credit (20/20),
        engaged but no points -> Partial Credit (10/20). Apply lenient floor:
        any submission with >= 1 graded message -> at least Partial Credit.

The script does NOT mutate stored DB scores; it just produces a CSV/Excel.
"""

import csv
import math
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
SRC = ROOT / "badm576_attempts.csv"
OUT_XLSX = ROOT / "BADM576_Lenient_Grades.xlsx"
OUT_CSV = ROOT / "BADM576_Lenient_Grades.csv"

TEST_EMAILS = {
    "ashishk@illinois.edu",
    "uiucbadm576@gmail.com",
    "student1@stakeholdersim.edu",
    "student2@stakeholdersim.edu",
    "student3@stakeholdersim.edu",
    "student4@stakeholdersim.edu",
    "student5@stakeholdersim.edu",
    "instructor@stakeholdersim.edu",
}

# Final score caps for each item
ITEM_MAX = {
    "classification": 20,   # LLM-graded HW quiz
    "pca": 10,              # MCQ HW check
    "ml_exercise": 20,      # Tier-mapped from AI score 0-100
}

ITEM_LABELS = {
    "classification": "Classification HW Quiz",
    "pca": "PCA HW Check",
    "ml_exercise": "ML Process Exercise",
}

ITEM_FULL_NAMES = {
    "classification": "BADM 576 — Classification HW Quiz",
    "pca": "BADM 576 — PCA HW Check",
    "ml_exercise": "BADM 576 — Week 7 ML Process In-Class Exercise",
}

ITEM_ADMIN_DATE = {
    "classification": "2026-04-22",
    "pca": "2026-04-29",
    "ml_exercise": "2026-03-11",
}


def netid_of(email: str) -> str:
    e = email.strip().lower()
    if e.endswith("@illinois.edu") or e.endswith("@illinois"):
        return e.split("@")[0]
    return e


def lenient_quiz_score(raw: float, max_score: float) -> float:
    """+20% engagement bump, capped at max, rounded up to nearest int."""
    bumped = min(max_score, raw + 0.20 * max_score)
    return math.ceil(bumped) if bumped > 0 else 0


def lenient_ml_exercise(ai_score: float) -> float:
    """Tier-mapped: AI score >0 -> 20 (full), engaged but 0 -> 10 (partial)."""
    if ai_score is None:
        return 0
    if ai_score > 0:
        return 20
    return 10  # showed up but earned no rubric points -> partial credit


def parse_date(s: str) -> datetime:
    # postgres dumps as "2026-04-22 20:50:17.500885"
    s = (s or "").split(".")[0].strip()
    return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")


def main():
    rows = []
    with SRC.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            email = (r["email"] or "").strip().lower()
            if not email or email in TEST_EMAILS:
                continue
            try:
                raw = float(r["score"] or 0)
                max_score = float(r["max_score"] or 0)
            except ValueError:
                continue
            try:
                dt = parse_date(r["submitted_at"])
            except Exception:
                continue
            rows.append({
                "item": r["item"],
                "email": email,
                "name": (r["name"] or "").strip(),
                "raw": raw,
                "max_raw": max_score,
                "date": dt,
            })

    # Build name -> canonical netid map from illinois.edu emails (so we can
    # merge personal-email signups onto the right NetID).
    name_to_netid = {}
    for r in rows:
        nid = netid_of(r["email"])
        if "@" not in nid:  # got a clean netid
            key = r["name"].strip().lower()
            if key and key not in name_to_netid:
                name_to_netid[key] = nid

    def canonical_netid(row) -> str:
        nid = netid_of(row["email"])
        if "@" not in nid:
            return nid
        name = row["name"].strip().lower()
        if not name:
            return nid
        # Exact match
        mapped = name_to_netid.get(name)
        if mapped:
            return mapped
        # Only merge when one name is a strict prefix/suffix subset of the other,
        # or one is a single token that exactly equals the other's first OR last
        # name AND there's exactly one such match in the population (otherwise
        # common last names like "zhang" cause false merges).
        tokens = name.split()
        candidates = []
        for stored_name, stored_nid in name_to_netid.items():
            stored_tokens = stored_name.split()
            if not stored_tokens or not tokens:
                continue
            if stored_name == name:
                candidates.append(stored_nid)
                continue
            # One side is a single token that equals the other's first OR last
            if len(tokens) == 1 and (tokens[0] == stored_tokens[0] or tokens[0] == stored_tokens[-1]):
                candidates.append(stored_nid)
            elif len(stored_tokens) == 1 and (stored_tokens[0] == tokens[0] or stored_tokens[0] == tokens[-1]):
                candidates.append(stored_nid)
        if len(candidates) == 1:
            return candidates[0]
        return nid

    # Best-per-(student, item) by raw score
    best = {}  # (item, netid) -> row
    for r in rows:
        nid = canonical_netid(r)
        k = (r["item"], nid)
        cur = best.get(k)
        if cur is None or r["raw"] > cur["raw"]:
            r = dict(r)
            r["netid"] = nid
            best[k] = r

    # Compute lenient scores
    for k, r in best.items():
        item = r["item"]
        max_final = ITEM_MAX[item]
        if item == "ml_exercise":
            r["lenient"] = lenient_ml_exercise(r["raw"])
        else:
            # quiz raw is on its own max -> rescale into the lenient bump
            r["lenient"] = lenient_quiz_score(r["raw"], max_final)
        r["max_final"] = max_final

    # Reshape: NetID -> {item -> row}
    by_student = defaultdict(dict)
    name_for = {}
    for (item, nid), r in best.items():
        by_student[nid][item] = r
        existing = name_for.get(nid)
        # Prefer multi-token "First Last" names over a single-word name/handle
        if existing is None or (" " in r["name"] and " " not in existing):
            name_for[nid] = r["name"]

    # ----- Build Excel -----
    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")

    def style_header(ws, row, n_cols):
        for col in range(1, n_cols + 1):
            c = ws.cell(row=row, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center")

    # --- Combined sheet ---
    ws = wb.create_sheet("All Students")

    title_font = Font(bold=True, size=12)
    subtitle_font = Font(italic=True, size=10, color="555555")

    # Row 1: course title
    ws.cell(row=1, column=1, value="BADM 576 — Lenient Grade Report (best attempt per student)").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    # Rows 2-4: per-item assignment full name + administered date
    item_meta_rows = [
        ("classification", "Classification HW Quiz", "C:D"),
        ("pca", "PCA HW Check", "E:F"),
        ("ml_exercise", "ML Process Exercise", "G:H"),
    ]
    for r_idx, (item, _, span) in enumerate(item_meta_rows, start=2):
        text = f"{ITEM_FULL_NAMES[item]}  ·  Administered {ITEM_ADMIN_DATE[item]}"
        ws.cell(row=r_idx, column=1, value=text).font = subtitle_font
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=9)

    HEADER_ROW = 6
    DATA_START = HEADER_ROW + 1

    headers = [
        "NetID", "Name",
        "Classification (/20)", "Classification Date",
        "PCA (/10)", "PCA Date",
        "ML Exercise (/20)", "ML Exercise Date",
        "Total (/50)",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=HEADER_ROW, column=col, value=h)
    style_header(ws, HEADER_ROW, len(headers))

    sorted_nids = sorted(by_student.keys())
    for i, nid in enumerate(sorted_nids, DATA_START):
        items = by_student[nid]
        c = items.get("classification")
        p = items.get("pca")
        m = items.get("ml_exercise")
        total = (c["lenient"] if c else 0) + (p["lenient"] if p else 0) + (m["lenient"] if m else 0)
        ws.cell(row=i, column=1, value=nid)
        ws.cell(row=i, column=2, value=name_for[nid])
        ws.cell(row=i, column=3, value=c["lenient"] if c else None)
        ws.cell(row=i, column=4, value=c["date"].strftime("%Y-%m-%d") if c else None)
        ws.cell(row=i, column=5, value=p["lenient"] if p else None)
        ws.cell(row=i, column=6, value=p["date"].strftime("%Y-%m-%d") if p else None)
        ws.cell(row=i, column=7, value=m["lenient"] if m else None)
        ws.cell(row=i, column=8, value=m["date"].strftime("%Y-%m-%d") if m else None)
        ws.cell(row=i, column=9, value=total)

    widths = [16, 26, 18, 16, 12, 16, 18, 16, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=DATA_START, column=1).coordinate
    last_data_row = DATA_START + len(sorted_nids) - 1
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(headers))}{last_data_row}"

    # --- Per-item sheets ---
    for item in ("classification", "pca", "ml_exercise"):
        sheet = wb.create_sheet(ITEM_LABELS[item])
        max_final = ITEM_MAX[item]
        sheet.cell(row=1, column=1, value=ITEM_FULL_NAMES[item]).font = Font(bold=True, size=12)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        sheet.cell(row=2, column=1, value=f"Administered {ITEM_ADMIN_DATE[item]}  ·  Best attempt per student, lenient scoring").font = Font(italic=True, size=10, color="555555")
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

        HDR = 4
        hdr = ["NetID", "Name", f"Lenient (/{max_final})", "Raw", "Max", "Date"]
        for col, h in enumerate(hdr, 1):
            sheet.cell(row=HDR, column=col, value=h)
        style_header(sheet, HDR, len(hdr))
        item_rows = sorted(
            [r for (it, _), r in best.items() if it == item],
            key=lambda r: r["netid"],
        )
        for i, r in enumerate(item_rows, HDR + 1):
            sheet.cell(row=i, column=1, value=r["netid"])
            sheet.cell(row=i, column=2, value=r["name"])
            sheet.cell(row=i, column=3, value=r["lenient"])
            sheet.cell(row=i, column=4, value=r["raw"])
            sheet.cell(row=i, column=5, value=r["max_raw"])
            sheet.cell(row=i, column=6, value=r["date"].strftime("%Y-%m-%d %H:%M"))
        for i, w in enumerate([16, 26, 14, 10, 8, 18], 1):
            sheet.column_dimensions[get_column_letter(i)].width = w
        sheet.freeze_panes = sheet.cell(row=HDR + 1, column=1).coordinate
        sheet.auto_filter.ref = f"A{HDR}:F{HDR + len(item_rows)}"

    wb.save(OUT_XLSX)

    # ----- Build flat CSV -----
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "NetID", "Name",
            "Classification_Lenient", "Classification_Raw", "Classification_Date",
            "PCA_Lenient", "PCA_Raw", "PCA_Date",
            "MLExercise_Lenient", "MLExercise_AIScore", "MLExercise_Date",
            "Total_Lenient_50",
        ])
        for nid in sorted_nids:
            items = by_student[nid]
            c = items.get("classification")
            p = items.get("pca")
            m = items.get("ml_exercise")
            total = (c["lenient"] if c else 0) + (p["lenient"] if p else 0) + (m["lenient"] if m else 0)
            w.writerow([
                nid, name_for[nid],
                c["lenient"] if c else "",
                c["raw"] if c else "",
                c["date"].strftime("%Y-%m-%d") if c else "",
                p["lenient"] if p else "",
                p["raw"] if p else "",
                p["date"].strftime("%Y-%m-%d") if p else "",
                m["lenient"] if m else "",
                m["raw"] if m else "",
                m["date"].strftime("%Y-%m-%d") if m else "",
                total,
            ])

    # ----- Summary print -----
    print(f"Wrote: {OUT_XLSX}")
    print(f"Wrote: {OUT_CSV}")
    print(f"\nUnique students: {len(sorted_nids)}")
    for item in ("classification", "pca", "ml_exercise"):
        n = sum(1 for it, _ in best if it == item)
        avg_lenient = (
            sum(r["lenient"] for (it, _), r in best.items() if it == item) / n
            if n else 0
        )
        print(f"  {ITEM_LABELS[item]:<25} {n:>3} students   avg lenient {avg_lenient:.2f}/{ITEM_MAX[item]}")


if __name__ == "__main__":
    main()
