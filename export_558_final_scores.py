"""Build the final BADM 558 quiz score sheet (one row per student, best attempt, capped at 15).

Reads the full DB dump (badm558_attempts_full.csv) which covers all sections.
"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = Path(__file__).parent / "badm558_attempts_full.csv"
OUT = Path(__file__).parent / "BADM558_Final_Scores_AllSections.xlsx"

MAX_POINTS = 15.0
TEST_EMAILS = {"ashishk@illinois.edu", "uiucbadm576@gmail.com"}

rows = []
with SRC.open(newline="", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for r in reader:
        if not r.get("email"):
            continue
        email = r["email"].strip().lower()
        if email in TEST_EMAILS:
            continue
        try:
            score = float(r["score"]) if r["score"] else 0.0
            max_score = float(r["max_score"]) if r["max_score"] else 0.0
        except ValueError:
            continue
        try:
            qa = int(r["questions_answered"] or 0)
        except ValueError:
            qa = 0
        if qa == 0:
            continue
        rows.append({
            "name": r["name"].strip(),
            "email": email,
            "raw_score": score,
            "max_score": max_score,
            "submitted": r["submitted_at"],
            "needs_review": int(r.get("needs_review_count") or 0) > 0,
            "questions_answered": qa,
        })

# Best attempt per student (by email)
best = {}
for row in rows:
    key = row["email"]
    if key not in best or row["raw_score"] > best[key]["raw_score"]:
        best[key] = row

# Cap at 15 and compute final percentage
final = []
for row in sorted(best.values(), key=lambda x: x["name"].lower()):
    capped = min(row["raw_score"], MAX_POINTS)
    final.append({
        **row,
        "final_score": capped,
        "final_pct": round(capped / MAX_POINTS * 100, 1),
        "over_cap": row["raw_score"] > MAX_POINTS,
    })

# Build workbook
wb = Workbook()
ws = wb.active
ws.title = "BADM 558 Final Scores"

headers = [
    "Student Name",
    "Email",
    "Final Score (/15)",
    "Final %",
    "Raw Score",
    "Questions Answered",
    "Needs Review",
    "Submitted At",
    "Notes",
]

header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True, color="FFFFFF")
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for i, row in enumerate(final, 2):
    notes = []
    if row["over_cap"]:
        notes.append(f"Answered {row['questions_answered']} (capped from {row['raw_score']:g})")
    if row["needs_review"]:
        notes.append("LLM/auto-graded - review")
    ws.cell(row=i, column=1, value=row["name"])
    ws.cell(row=i, column=2, value=row["email"])
    ws.cell(row=i, column=3, value=row["final_score"])
    ws.cell(row=i, column=4, value=row["final_pct"])
    ws.cell(row=i, column=5, value=row["raw_score"])
    ws.cell(row=i, column=6, value=row["questions_answered"])
    ws.cell(row=i, column=7, value="Yes" if row["needs_review"] else "No")
    ws.cell(row=i, column=8, value=row["submitted"])
    ws.cell(row=i, column=9, value="; ".join(notes))

# Column widths
widths = [26, 32, 16, 10, 12, 18, 14, 30, 40]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(final)+1}"

# Summary sheet
summ = wb.create_sheet("Summary")
n = len(final)
total = sum(r["final_score"] for r in final)
mean = total / n if n else 0
median_vals = sorted(r["final_score"] for r in final)
median = median_vals[n // 2] if n % 2 else (median_vals[n // 2 - 1] + median_vals[n // 2]) / 2 if n else 0
perfect = sum(1 for r in final if r["final_score"] >= MAX_POINTS)
zero = sum(1 for r in final if r["final_score"] == 0)

summary_rows = [
    ("Quiz", "BADM 558 - Big Data Infrastructure Quiz Prep"),
    ("Date administered", "2026-03-31"),
    ("Max score", MAX_POINTS),
    ("Students (unique)", n),
    ("Mean", round(mean, 2)),
    ("Median", median),
    ("Perfect scores (15/15)", perfect),
    ("Zero scores", zero),
]
for i, (k, v) in enumerate(summary_rows, 1):
    ws2_a = summ.cell(row=i, column=1, value=k)
    ws2_a.font = Font(bold=True)
    summ.cell(row=i, column=2, value=v)
summ.column_dimensions["A"].width = 30
summ.column_dimensions["B"].width = 50

wb.save(OUT)
print(f"Wrote: {OUT}")
print(f"Students: {n}, Mean: {mean:.2f}/15, Median: {median}/15, Perfect: {perfect}, Zero: {zero}")
