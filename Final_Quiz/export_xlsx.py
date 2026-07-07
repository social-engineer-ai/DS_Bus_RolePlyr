"""Export the BADM 576 final-quiz scores to a formatted xlsx.

Reads `badm576_final_scores_after_manual.csv` and writes:
- A 'Scores' sheet with header row, conditional fill on the total column, and
  per-section subtotals.
- A 'Class Stats' sheet with min/max/mean/median per section and grade-bucket
  counts.

Saves to the local Final_Quiz folder AND the 576_grading/Final_Quiz folder.
"""

import csv
import statistics
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

SRC_CSV = Path("badm576_final_scores_after_manual.csv")
OUT_LOCAL = Path("BADM576_Final_Scores.xlsx")
OUT_SHARED = Path(
    r"C:\Users\ashishk\Dropbox\My PC (BUS-P10E67720)\Documents\Development"
    r"\576_grading\Final_Quiz\BADM576_Final_Scores.xlsx"
)


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # navy
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def load_rows():
    rows = []
    with open(SRC_CSV, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            rows.append({
                "netid": r["netid"],
                "name": r["name"],
                "submitted_at": r["submitted_at"][:16].replace("T", " "),
                "section_a": float(r["section_a_mcq_40"]),
                "section_b": float(r["section_b_short_40"]),
                "section_c": float(r["section_c_scenarios_20"]),
                "total": float(r["total_100"]),
            })
    return rows


def build_scores_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Scores"

    headers = [
        "Rank",
        "NetID",
        "Name",
        "Submitted (UTC)",
        "Section A — MCQ (/40)",
        "Section B — Short Answer (/40)",
        "Section C — Scenarios (/20)",
        "Total (/100)",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    rows_sorted = sorted(rows, key=lambda r: -r["total"])
    for i, r in enumerate(rows_sorted, start=2):
        ws.cell(row=i, column=1, value=i - 1).alignment = CENTER
        ws.cell(row=i, column=2, value=r["netid"]).alignment = LEFT
        ws.cell(row=i, column=3, value=r["name"]).alignment = LEFT
        ws.cell(row=i, column=4, value=r["submitted_at"]).alignment = CENTER
        ws.cell(row=i, column=5, value=r["section_a"]).alignment = CENTER
        ws.cell(row=i, column=6, value=r["section_b"]).alignment = CENTER
        ws.cell(row=i, column=7, value=r["section_c"]).alignment = CENTER
        total_cell = ws.cell(row=i, column=8, value=r["total"])
        total_cell.alignment = CENTER
        total_cell.fill = TOTAL_FILL
        total_cell.font = Font(bold=True)

        for col in range(1, 9):
            ws.cell(row=i, column=col).border = BORDER

    last_row = len(rows_sorted) + 1

    # Conditional formatting on Total column (col H)
    total_range = f"H2:H{last_row}"
    ws.conditional_formatting.add(
        total_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["85"],
                   fill=PatternFill("solid", fgColor="C6EFCE"),  # green
                   font=Font(bold=True, color="006100")),
    )
    ws.conditional_formatting.add(
        total_range,
        CellIsRule(operator="between", formula=["60", "84.999"],
                   fill=PatternFill("solid", fgColor="FFEB9C"),  # yellow
                   font=Font(bold=True, color="9C5700")),
    )
    ws.conditional_formatting.add(
        total_range,
        CellIsRule(operator="lessThan", formula=["60"],
                   fill=PatternFill("solid", fgColor="FFC7CE"),  # red
                   font=Font(bold=True, color="9C0006")),
    )

    # Column widths
    widths = [6, 24, 26, 18, 22, 30, 24, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 38

    # Header row alignment with wrap
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"


def build_stats_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Class Stats")

    sec_a = [r["section_a"] for r in rows]
    sec_b = [r["section_b"] for r in rows]
    sec_c = [r["section_c"] for r in rows]
    tot = [r["total"] for r in rows]

    headers = ["Metric", "Section A (/40)", "Section B (/40)", "Section C (/20)", "Total (/100)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    def stats_row(label, vals):
        return [label, round(min(vals), 2), round(max(vals), 2),
                round(statistics.mean(vals), 2), round(statistics.median(vals), 2),
                round(statistics.stdev(vals), 2) if len(vals) > 1 else 0]

    # Use a different layout: rows for each section, columns for each metric.
    ws.delete_rows(1)
    ws.append(["Statistic", "Section A (/40)", "Section B (/40)", "Section C (/20)", "Total (/100)"])
    rows_to_add = [
        ("Min", min(sec_a), min(sec_b), min(sec_c), min(tot)),
        ("Max", max(sec_a), max(sec_b), max(sec_c), max(tot)),
        ("Mean", round(statistics.mean(sec_a), 2), round(statistics.mean(sec_b), 2),
         round(statistics.mean(sec_c), 2), round(statistics.mean(tot), 2)),
        ("Median", round(statistics.median(sec_a), 2), round(statistics.median(sec_b), 2),
         round(statistics.median(sec_c), 2), round(statistics.median(tot), 2)),
        ("Std dev", round(statistics.stdev(sec_a), 2), round(statistics.stdev(sec_b), 2),
         round(statistics.stdev(sec_c), 2), round(statistics.stdev(tot), 2)),
    ]
    for r in rows_to_add:
        ws.append(r)

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for r in range(2, 7):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = CENTER if c > 1 else LEFT
            if c == 1:
                cell.font = Font(bold=True)

    # Grade buckets
    ws.append([])
    ws.append(["Grade buckets (Total)", "Count", "Students"])
    for cell in ws[ws.max_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    bucket_defs = [
        ("90-100 (A)", lambda x: x >= 90),
        ("80-89 (B)", lambda x: 80 <= x < 90),
        ("70-79 (C)", lambda x: 70 <= x < 80),
        ("60-69 (D)", lambda x: 60 <= x < 70),
        ("Below 60 (F)", lambda x: x < 60),
    ]
    bucket_fill = {
        "90-100 (A)": PatternFill("solid", fgColor="C6EFCE"),
        "80-89 (B)": PatternFill("solid", fgColor="DDEBF7"),
        "70-79 (C)": PatternFill("solid", fgColor="FFEB9C"),
        "60-69 (D)": PatternFill("solid", fgColor="FCE4D6"),
        "Below 60 (F)": PatternFill("solid", fgColor="FFC7CE"),
    }
    rows_sorted_by_total_desc = sorted(rows, key=lambda r: -r["total"])
    for label, pred in bucket_defs:
        names = [r["name"] for r in rows_sorted_by_total_desc if pred(r["total"])]
        ws.append([label, len(names), ", ".join(names)])
        for cell in ws[ws.max_row]:
            cell.border = BORDER
            cell.alignment = LEFT if cell.column == 1 or cell.column == 3 else CENTER
            cell.fill = bucket_fill[label]
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 80
    for col in ("D", "E"):
        ws.column_dimensions[col].width = 18

    ws.freeze_panes = "A2"


def main():
    rows = load_rows()
    wb = Workbook()
    build_scores_sheet(wb, rows)
    build_stats_sheet(wb, rows)
    wb.save(OUT_LOCAL)
    print(f"Saved: {OUT_LOCAL.resolve()}")
    wb.save(OUT_SHARED)
    print(f"Saved: {OUT_SHARED}")


if __name__ == "__main__":
    main()
