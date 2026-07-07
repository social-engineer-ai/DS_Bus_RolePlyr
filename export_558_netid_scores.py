"""Final BADM 558 score sheet: Net ID + Score, one row per student (best attempt, capped at 15)."""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = Path(__file__).parent / "badm558_attempts_full.csv"
OUT = Path(__file__).parent / "BADM558_NetID_Scores.xlsx"

MAX_POINTS = 15.0
TEST_EMAILS = {"ashishk@illinois.edu", "uiucbadm576@gmail.com"}

rows = []
with SRC.open(newline="", encoding="utf-8") as f:
    for r in csv.DictReader(f):
        email = r["email"].strip().lower()
        if email in TEST_EMAILS:
            continue
        try:
            score = float(r["score"] or 0)
            qa = int(r["questions_answered"] or 0)
        except ValueError:
            continue
        if qa == 0:
            continue
        rows.append({
            "name": r["name"].strip(),
            "email": email,
            "score": score,
        })

# Best per student
best = {}
for row in rows:
    e = row["email"]
    if e not in best or row["score"] > best[e]["score"]:
        best[e] = row

# Build records: derive Net ID from email
records = []
for row in best.values():
    email = row["email"]
    if email.endswith("@illinois.edu"):
        netid = email.split("@")[0]
    else:
        netid = email
    records.append({
        "name": row["name"],
        "netid": netid,
        "email": email,
        "score": min(row["score"], MAX_POINTS),
    })

records.sort(key=lambda r: r["netid"])

# Build workbook
wb = Workbook()
ws = wb.active
ws.title = "BADM 558 Final Scores"

headers = ["Net ID", "Student Name", "Score (/15)"]
header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True, color="FFFFFF")
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal="center")

for i, r in enumerate(records, 2):
    ws.cell(row=i, column=1, value=r["netid"])
    ws.cell(row=i, column=2, value=r["name"])
    ws.cell(row=i, column=3, value=r["score"])

ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 14
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:C{len(records)+1}"

wb.save(OUT)
print(f"Wrote: {OUT}")
print(f"Students: {len(records)}")
